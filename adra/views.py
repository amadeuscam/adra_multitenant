import os
import time
from datetime import date
import telegram
from PyPDF2 import PdfFileReader, PdfFileWriter
from PyPDF2.generic import BooleanObject, NameObject, IndirectObject
from allauth.account.adapter import DefaultAccountAdapter
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Q
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import (ListView, DetailView, DeleteView, UpdateView, CreateView)
from mailmerge import MailMerge
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from rest_framework import permissions
from rest_framework import viewsets
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail)
# from .tasks import export_zip, delete_folder
from tenant_schemas.utils import remove_www
from .filters import AlimentosFilters
from .forms import AlimentosFrom, HijoForm, PersonaForm, UserEditForm, GastoForm
from .models import Persona, Alimentos, AlmacenAlimentos, Hijo, DelegacionData, Gasto
from .serializers import PersonaSerializer, UserSerializer


# from django.views.decorators.cache import cache_page

# logger = logging.getLogger(__name__)


def anuncios(request):
    return render(request, 'adra/anuncio.html')


@login_required
def edit(request):
    if request.method == 'POST':
        user_form = UserEditForm(instance=request.user, data=request.POST)
        # profile_form = ProfileEditForm(
        #     instance=request.user.profile,
        #     data=request.POST,
        #     files=request.FILES)
        # and profile_form.is_valid()
        if user_form.is_valid():
            user_form.save()
            # profile_form.save()
            messages.success(
                request, 'El perfil se ha actualizado correctamente')
            return redirect('adra:edit-profile')
        else:
            messages.error(request, 'Error updating your profile')
    else:
        user_form = UserEditForm(instance=request.user)
        # profile_form = ProfileEditForm(instance=request.user.profile)
    return render(request,
                  'account/edit-perfile.html',
                  {'user_form': user_form})


class PersonaListView(LoginRequiredMixin, ListView):
    template_name = 'adra/index.html'
    context_object_name = 'ultima_persona'
    model = Persona
    paginate_by = 15
    login_url = 'account_login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # enviar al misma pagina varios contextos
        context['hijo'] = Hijo.objects
        context['hijomenor'] = Hijo
        context['nbar'] = "home"
        dels = DelegacionData.objects.first()
        self.request.session['platform_name'] = str(self.request.tenant)
        self.request.session['telegram_active'] = dels.telegram_code

        return context


# ir a la pagina detallada de cada persona
class PersonaDetailView(LoginRequiredMixin, DetailView):
    model = Persona
    template_name = 'adra/detail.html'
    context_object_name = 'persona'
    login_url = 'account_login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = AlimentosFilters(self.request.GET,
                                             queryset=self.get_queryset())
        return context


class BuscarDetailView(LoginRequiredMixin, ListView):
    model = Alimentos
    template_name = 'busqueda_a/view.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = AlimentosFilters(self.request.GET,
                                             queryset=self.get_queryset())
        return context


class PersonaCreateView(LoginRequiredMixin, CreateView):
    model = Persona
    success_url = '/'
    form_class = PersonaForm

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        messages.add_message(self.request, messages.SUCCESS, f'Beneficiario añadido correctamente!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            p = Persona.objects.latest('created_at')
            context['nbar'] = "create"
            context['bas'] = p
            return context
        except Persona.DoesNotExist:
            pass


class PersonaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Persona
    fields = [
        'nombre_apellido',
        'dni',
        'otros_documentos',
        'fecha_nacimiento',
        'numero_adra',
        'nacionalidad',
        'domicilio',
        'are_acte',
        'ciudad',
        'telefono',
        'mensaje',
        'sexo',
        'discapacidad',
        'domingo',
        'empadronamiento',
        'libro_familia',
        'fotocopia_dni',
        'prestaciones',
        'nomnia',
        'cert_negativo',
        'aquiler_hipoteca',
        'recibos',
        'email',
        'covid'
    ]
    success_message = 'Datele sau salvat cu success!!'

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['up'] = "update"
        return context


class PersonaDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Persona
    success_message = 'Beneficirio borrado corectamente'
    success_url = '/'


def adauga_alimentos_persona(request, pk):
    persona = get_object_or_404(Persona, pk=pk)
    almacen = AlmacenAlimentos.objects.get(pk=1)

    if request.method == 'POST':
        a_form = AlimentosFrom(request.POST)
        if a_form.is_valid():
            alimentos = a_form.save(commit=False)

            almacen.alimento_1 -= alimentos.alimento_1
            almacen.alimento_2 -= alimentos.alimento_2
            almacen.alimento_3 -= alimentos.alimento_3
            almacen.alimento_4 -= alimentos.alimento_4
            # almacen.alimento_5 -= alimentos.alimento_5
            almacen.alimento_6 -= alimentos.alimento_6
            almacen.alimento_7 -= alimentos.alimento_7
            almacen.alimento_8 -= alimentos.alimento_8
            almacen.alimento_9 -= alimentos.alimento_9
            almacen.alimento_10 -= alimentos.alimento_10
            almacen.alimento_11 -= alimentos.alimento_11
            almacen.alimento_12 -= alimentos.alimento_12
            almacen.alimento_13 -= alimentos.alimento_13
            almacen.alimento_14 -= alimentos.alimento_14
            almacen.alimento_15 -= alimentos.alimento_15

            alimentos.persona = persona
            alimentos.modificado_por = request.user
            almacen.save()
            alimentos.save()
            # logger.warning("aliemnte sau dat beneficiarului")
            messages.success(request, 'Los alimentos se han entregado correctamente!')
            return redirect(persona)

    else:
        a_form = AlimentosFrom()
    return render(request, 'adra/alimentos_form.html', {'form': a_form})


class PersonaAlimentosUpdateView(LoginRequiredMixin, UpdateView):
    model = Alimentos
    fields = [
        'alimento_1',
        'alimento_2',
        'alimento_3',
        'alimento_4',
        'alimento_6',
        'alimento_7',
        'alimento_8',
        'alimento_9',
        'alimento_10',
        'alimento_11',
        'alimento_12',
        'alimento_13',
        'alimento_14',
        'alimento_15',
        'fecha_recogida'

    ]

    def form_valid(self, form):
        clean = form.changed_data
        almacen = AlmacenAlimentos.objects.get(pk=1)

        if "alimento_1" in clean:
            valor_anterior_alimento_1 = form.initial["alimento_1"]
            if form.instance.alimento_1 > valor_anterior_alimento_1:
                restante = abs(form.instance.alimento_1 - valor_anterior_alimento_1)
                almacen.alimento_1 -= restante
            else:
                restante = abs(form.instance.alimento_1 - valor_anterior_alimento_1)
                almacen.alimento_1 += restante

        if "alimento_2" in clean:
            valor_anterior_alimento_2 = form.initial["alimento_2"]
            if form.instance.alimento_2 > valor_anterior_alimento_2:
                restante = abs(form.instance.alimento_2 - valor_anterior_alimento_2)
                almacen.alimento_2 -= restante
            else:
                restante = abs(form.instance.alimento_2 - valor_anterior_alimento_2)
                almacen.alimento_2 += restante

        if "alimento_3" in clean:
            valor_anterior_alimento_3 = form.initial["alimento_3"]
            if form.instance.alimento_3 > valor_anterior_alimento_3:
                restante = abs(form.instance.alimento_3 - valor_anterior_alimento_3)
                almacen.alimento_3 -= restante
            else:
                restante = abs(form.instance.alimento_3 - valor_anterior_alimento_3)
                almacen.alimento_3 += restante

        if "alimento_4" in clean:
            valor_anterior_alimento_4 = form.initial["alimento_4"]
            if form.instance.alimento_4 > valor_anterior_alimento_4:
                restante = abs(form.instance.alimento_4 - valor_anterior_alimento_4)
                almacen.alimento_4 -= restante
            else:
                restante = abs(form.instance.alimento_4 - valor_anterior_alimento_4)
                almacen.alimento_4 += restante

        if "alimento_6" in clean:
            valor_anterior_alimento_6 = form.initial["alimento_6"]
            if form.instance.alimento_6 > valor_anterior_alimento_6:
                restante = abs(form.instance.alimento_6 - valor_anterior_alimento_6)
                almacen.alimento_6 -= restante
            else:
                restante = abs(form.instance.alimento_6 - valor_anterior_alimento_6)
                almacen.alimento_6 += restante

        if "alimento_7" in clean:
            valor_anterior_alimento_7 = form.initial["alimento_7"]
            if form.instance.alimento_7 > valor_anterior_alimento_7:
                restante = abs(form.instance.alimento_7 - valor_anterior_alimento_7)
                almacen.alimento_7 -= restante
            else:
                restante = abs(form.instance.alimento_7 - valor_anterior_alimento_7)
                almacen.alimento_7 += restante

        if "alimento_8" in clean:
            valor_anterior_alimento_8 = form.initial["alimento_8"]
            if form.instance.alimento_8 > valor_anterior_alimento_8:
                restante = abs(form.instance.alimento_8 - valor_anterior_alimento_8)
                almacen.alimento_8 -= restante
            else:
                restante = abs(form.instance.alimento_8 - valor_anterior_alimento_8)
                almacen.alimento_8 += restante

        if "alimento_9" in clean:
            valor_anterior_alimento_9 = form.initial["alimento_9"]
            if form.instance.alimento_9 > valor_anterior_alimento_9:
                restante = abs(form.instance.alimento_9 - valor_anterior_alimento_9)
                almacen.alimento_9 -= restante
            else:
                restante = abs(form.instance.alimento_9 - valor_anterior_alimento_9)
                almacen.alimento_9 += restante

        if "alimento_10" in clean:
            valor_anterior_alimento_10 = form.initial["alimento_10"]
            if form.instance.alimento_10 > valor_anterior_alimento_10:
                restante = abs(form.instance.alimento_10 - valor_anterior_alimento_10)
                almacen.alimento_10 -= restante
            else:
                restante = abs(form.instance.alimento_10 - valor_anterior_alimento_10)
                almacen.alimento_10 += restante

        if "alimento_11" in clean:
            valor_anterior_alimento_11 = form.initial["alimento_11"]
            if form.instance.alimento_11 > valor_anterior_alimento_11:
                restante = abs(form.instance.alimento_11 - valor_anterior_alimento_11)
                almacen.alimento_11 -= restante
            else:
                restante = abs(form.instance.alimento_11 - valor_anterior_alimento_11)
                almacen.alimento_11 += restante

        if "alimento_12" in clean:
            valor_anterior_alimento_12 = form.initial["alimento_12"]
            if form.instance.alimento_12 > valor_anterior_alimento_12:
                restante = abs(form.instance.tarito_fruta - valor_anterior_alimento_12)
                almacen.alimento_12 -= restante
            else:
                restante = abs(form.instance.tarito_fruta - valor_anterior_alimento_12)
                almacen.alimento_12 += restante

        if "alimento_13" in clean:
            valor_anterior_alimento_13 = form.initial["alimento_13"]
            if form.instance.alimento_13 > valor_anterior_alimento_13:
                restante = abs(form.instance.alimento_13 - valor_anterior_alimento_13)
                almacen.alimento_13 -= restante
            else:
                restante = abs(form.instance.alimento_13 - valor_anterior_alimento_13)
                almacen.alimento_13 += restante

        if "alimento_14" in clean:
            valor_anterior_alimento_14 = form.initial["alimento_14"]
            if form.instance.alimento_14 > valor_anterior_alimento_14:
                restante = abs(form.instance.alimento_14 - valor_anterior_alimento_14)
                almacen.alimento_14 -= restante
            else:
                restante = abs(form.instance.alimento_14 - valor_anterior_alimento_14)
                almacen.alimento_14 += restante

        if "alimento_15" in clean:
            valor_anterior_alimento_15 = form.initial["alimento_15"]
            if form.instance.alimento_15 > valor_anterior_alimento_15:
                restante = abs(form.instance.alimento_15 - valor_anterior_alimento_15)
                almacen.alimento_15 -= restante
            else:
                restante = abs(form.instance.alimento_15 - valor_anterior_alimento_15)
                almacen.alimento_15 += restante

        almacen.save()
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)


class PersonaAlimentosDeleteView(LoginRequiredMixin, DeleteView):
    model = Alimentos
    success_url = '/'

    def test_func(self):
        persona = self.get_object()
        if self.request.user == persona.modificado_por:
            return True
        return False


class AlmacenListView(LoginRequiredMixin, ListView):
    model = AlmacenAlimentos
    template_name = 'adra_almacen/index.html'
    context_object_name = 'almacen_adra'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nbar'] = "almacen_a"
        return context


class HijoCreateView(LoginRequiredMixin, CreateView):
    model = Hijo
    form_class = HijoForm

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)


def adauga_hijo_persona(request, pk):
    persona = get_object_or_404(Persona, pk=pk)
    if request.method == 'POST':
        h_form = HijoForm(request.POST)
        if h_form.is_valid():
            hijo = h_form.save(commit=False)
            hijo.persona = persona
            hijo.modificado_por = request.user
            hijo.save()
            return redirect(persona)

    else:
        h_form = HijoForm()
    return render(request, 'adra/hijo_form.html', {'form': h_form})


class HijoUpdateView(LoginRequiredMixin, UpdateView):
    model = Hijo
    fields = [
        'parentesco',
        'nombre_apellido',
        'dni',
        'otros_documentos',
        'fecha_nacimiento',

    ]

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)

    def test_func(self):
        persona = self.get_object()
        if self.request.user == persona.modificado_por:
            return True
        return


class HijoDeleteView(LoginRequiredMixin, DeleteView):
    model = Hijo
    persoan = Persona
    success_url = reverse_lazy("adra:persona-home")


@login_required
def buscar(request):
    if 'q' in request.GET and request.GET['q']:
        q = request.GET['q']

        if (q.isdigit()):
            ultima_persona = Persona.objects.filter(
                Q(numero_adra=q) | Q(telefono=q)).filter(active=True)
        else:
            ultima_persona = Persona.objects.filter(
                nombre_apellido__icontains=q).filter(active=True)
        if ultima_persona:
            return render(request, 'adra/index.html',
                          {'ultima_persona': ultima_persona, 'query': q})
        else:
            beneficiarios_no_activos = Persona.objects.filter(active=False).order_by("-numero_adra")
            return render(request, 'adra/no_beneficiarios.html', {'ben': beneficiarios_no_activos})
    else:
        return HttpResponse('<h1>Por favor buscar con otro criterio</h1>')


def calculate_age(age):
    today = date.today()
    return today.year - age.year - ((today.month, today.day) < (age.month, age.day))


@login_required
def statistics_persona(request):
    mod = 2
    if request.POST:
        # if request.POST.get('fecha_val', None):
        #     fecha_val = request.POST.get('fecha_val', None)
        #     rep = export_zip(fecha_val)
        #     delete_folder.apply_async()
        #     return rep

        if request.POST.get('mod', None):
            mod = request.POST.get('mod', None)

    if int(mod) == 1:
        beneficiar = Persona.objects.prefetch_related('hijo').filter(active=True).exclude(covid=True)
    else:
        beneficiar = Persona.objects.prefetch_related('hijo').filter(active=True)

    tic = time.perf_counter()

    lst_02_mujer = []
    lst_02_hombre = []

    lst_3_15_mujer = []
    lst_3_15_hombre = []

    lst_16_64_mujer = []
    lst_16_64_hombre = []

    lst_64_mujer = []
    lst_64_hombre = []

    [lst_02_mujer.append(b.age) for b in beneficiar if 0 <= b.age <= 2 and b.sexo == "mujer"]
    [lst_02_hombre.append(b.age) for b in beneficiar if 0 <= b.age <= 2 and b.sexo == "hombre"]

    [lst_3_15_mujer.append(b.age) for b in beneficiar if 3 <= b.age <= 15 and b.sexo == "mujer"]
    [lst_3_15_hombre.append(b.age) for b in beneficiar if 3 <= b.age <= 15 and b.sexo == "hombre"]

    [lst_16_64_mujer.append(b.age) for b in beneficiar if 16 <= b.age <= 64 and b.sexo == "mujer"]
    [lst_16_64_hombre.append(b.age) for b in beneficiar if 16 <= b.age <= 64 and b.sexo == "hombre"]

    [lst_64_mujer.append(b.age) for b in beneficiar if b.age >= 65 and b.sexo == "mujer"]
    [lst_64_hombre.append(b.age) for b in beneficiar if b.age >= 65 and b.sexo == "hombre"]

    lst_familiares = []
    for hijo in beneficiar:
        [(lst_02_mujer.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         0 <= b.age <= 2 and b.sexo == "m"]
        [(lst_02_hombre.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         0 <= b.age <= 2 and b.sexo == "h"]

        [(lst_3_15_mujer.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         3 <= b.age <= 15 and b.sexo == "m"]
        [(lst_3_15_hombre.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         3 <= b.age <= 15 and b.sexo == "h"]

        [(lst_16_64_mujer.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         16 <= b.age <= 64 and b.sexo == "m"]
        [(lst_16_64_hombre.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         16 <= b.age <= 64 and b.sexo == "h"]

        [(lst_64_mujer.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         b.age >= 65 and b.sexo == "m"]
        [(lst_64_hombre.append(b.age), lst_familiares.append(b.age)) for b in hijo.hijo.all() if
         b.age >= 65 and b.sexo == "h"]

    data_statistics = {
        "total_per_mujer_02": len(lst_02_mujer),
        "total_per_mujer_03": len(lst_3_15_mujer),
        "total_per_mujer_16": len(lst_16_64_mujer),
        "total_per_mujer_65": len(lst_64_mujer),
        "total_mujeres": len(lst_02_mujer) + len(lst_3_15_mujer) + len(lst_16_64_mujer) + len(lst_64_mujer),

        "total_per_hombre_02": len(lst_02_hombre),
        "total_per_hombre_03": len(lst_3_15_hombre),
        "total_per_hombre_16": len(lst_16_64_hombre),
        "total_per_hombre_65": len(lst_64_hombre),
        "total_hombres": len(lst_02_hombre) + len(lst_3_15_hombre) + len(lst_16_64_hombre) + len(lst_64_hombre),
        "total_02": len(lst_02_mujer) + len(lst_02_hombre),
        "total_03": len(lst_3_15_mujer) + len(lst_3_15_hombre),
        "total_16": len(lst_16_64_mujer) + len(lst_16_64_hombre),
        "total_65": len(lst_64_mujer) + len(lst_64_hombre),
        "total_personas": len(lst_02_mujer) + len(lst_02_hombre) + len(lst_3_15_mujer) + len(lst_3_15_hombre) + len(
            lst_16_64_mujer) + len(lst_16_64_hombre) + len(lst_64_mujer) + len(lst_64_hombre),
        "discapacidad": Persona.objects.filter(discapacidad=True, active=True).count(),
        "total_beneficiarios": beneficiar.count(),
        "total_familiares": len(lst_familiares),
        "nbar": 'stat',
        "mod": mod

    }
    toc = time.perf_counter()
    print(f"Load page in {toc - tic:0.4f} seconds")

    return render(request, 'statistics/index.html', data_statistics)


def telegram_messages(request):
    print(request.POST.get('dom_select'))
    dom = request.POST.get('dom_select')
    mensaje_propio = request.POST.get('mensaje_propio')

    if dom and mensaje_propio:
        bot = telegram.Bot(token=settings.TELEGRAM_TOKEN_KEY)
        persona = Persona.objects.filter(active=True).filter(Q(domingo=f"Domingo {int(dom)}") | Q(domingo=int(dom)),
                                                             ciudad__icontains="Torrejon de ardoz").exclude(covid=True)
        per_list = [p.nombre_apellido for p in persona]
        list_format = '\n'.join(per_list)
        bot.send_message('-1001438819726', f"*{list_format}*", parse_mode=telegram.ParseMode.MARKDOWN_V2)
        bot.send_message('-1001438819726', f"*{mensaje_propio}*", parse_mode=telegram.ParseMode.MARKDOWN_V2)
        messages.success(request, 'El mensaje se ha mandado correctamente ')

    return render(request, 'telegram/index.html', {'nbar': "tel"})


def gastos_delegacion(request):
    gastos = Gasto.objects.all()

    if request.method == 'POST':
        g_form = GastoForm(request.POST)
        if g_form.is_valid():
            gasto = g_form.save(commit=False)
            gasto.save()
            return redirect('adra:gastos_delegacion_view')

    else:
        g_form = GastoForm()
    return render(request, 'gastos/index.html', {'nbar': "gast", 'gasto_ret': gastos, "form_g":g_form})


def export_users_csv(request):
    beneficiarios_queryset = Persona.objects.order_by('numero_adra').filter(active=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=beneficiarios.xlsx'
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    # Delete the default worksheet
    # workbook.remove(workbook.active)
    worksheet.title = 'Beneficiarios'
    # Define the titles for columns
    worksheet.column_dimensions['B'].width = 40
    worksheet.column_dimensions['C'].width = 40
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['E'].width = 40
    columns = [
        'Numar adra',
        'Nombre',
        'Representante familiar',
        'Dni',
        'Pasaporte',
        'Fecha de nacimiento',
    ]
    row_num = 1
    fill = PatternFill(start_color='43fb32', fill_type='solid')
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    count = 0
    for ben in beneficiarios_queryset:
        row_num += 1
        count += 1

        # Define the data for each cell in the row
        row = [
            f"{count}-{ben.numero_adra}",
            ben.nombre_apellido,
            1,
            ben.dni,
            ben.otros_documentos,
            ben.fecha_nacimiento.strftime("%d/%m/%Y"),
        ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

        for d in ben.hijo.filter(active=True):
            count += 1
            row_hijos = [
                f'{count}',
                d.nombre_apellido,
                '',
                d.dni,
                d.otros_documentos,
                d.fecha_nacimiento.strftime("%d/%m/%Y"),
            ]
            row_num += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row_hijos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal='center')

    workbook.save(response)

    return response


@login_required
def buscar_fecha(request):
    alimentos_list = Alimentos.objects.all()
    user_filter = AlimentosFilters(request.GET, queryset=alimentos_list)
    # EXCEL ZONE
    # # Create a workbook and add a worksheet.
    # workbook = xlsxwriter.Workbook('AlimentosEnero-Julio.xlsx')
    # worksheet = workbook.add_worksheet()
    # # Start from the first cell. Rows and columns are zero indexed.
    # row = 0
    # col = 0
    # bold = workbook.add_format({'bold': True})
    #
    #
    # for u in user_filter.qs:
    #     worksheet.write(row, col, f'{u.persona}')
    #     worksheet.write(row, col + 1, f'{u.fecha_recogida}')
    #     worksheet.write_number(row, col + 2, u.arroz_blanco)
    #     worksheet.write_number(row, col + 3, u.garbanzo_cocido)
    #     worksheet.write_number(row, col + 4, u.atun_sardina)
    #     worksheet.write_number(row, col + 5, u.sardina)
    #     worksheet.write_number(row, col + 6, u.pasta_espagueti)
    #     worksheet.write_number(row, col + 7, u.tomate_frito)
    #     worksheet.write_number(row, col + 8, u.galletas)
    #     worksheet.write_number(row, col + 9, u.macedonia_verdura_conserva)
    #     worksheet.write_number(row, col + 10, u.fruta_conserva_pera)
    #     worksheet.write_number(row, col + 11, u.fruta_conserva_coctel)
    #     worksheet.write_number(row, col + 12, u.tarito_pollo)
    #     worksheet.write_number(row, col + 13, u.tarito_fruta)
    #     worksheet.write_number(row, col + 14, u.leche)
    #     worksheet.write_number(row, col + 15, u.batido_chocolate)
    #     worksheet.write_number(row, col + 16, u.aceite_de_oliva)
    #     worksheet.write(row, col + 17, f'{u.modificado_por}')
    #     row += 1
    #
    #     print(u.fecha_recogida)
    # workbook.close()
    # print(arroz_sum)
    # alubia_sum = user_filter.qs.aggregate(Sum('alubia_blanca'))
    alimento_1 = user_filter.qs.aggregate(Sum('alimento_1'))
    alimento_2 = user_filter.qs.aggregate(Sum('alimento_2'))
    alimento_3 = user_filter.qs.aggregate(Sum('alimento_3'))
    alimento_4 = user_filter.qs.aggregate(Sum('alimento_4'))
    alimento_6 = user_filter.qs.aggregate(Sum('alimento_6'))
    alimento_7 = user_filter.qs.aggregate(Sum('alimento_7'))
    alimento_8 = user_filter.qs.aggregate(Sum('alimento_8'))
    alimento_9 = user_filter.qs.aggregate(Sum('alimento_9'))
    alimento_10 = user_filter.qs.aggregate(Sum('alimento_10'))
    alimento_11 = user_filter.qs.aggregate(Sum('alimento_11'))
    alimento_12 = user_filter.qs.aggregate(Sum('alimento_12'))
    alimento_13 = user_filter.qs.aggregate(Sum('alimento_13'))
    alimento_14 = user_filter.qs.aggregate(Sum('alimento_14'))
    alimento_15 = user_filter.qs.aggregate(Sum('alimento_15'))
    # aceite_de_oliva_sum = user_filter.qs.aggregate(Sum('aceite_de_oliva'))

    return render(request, 'busqueda_a/view.html',
                  {
                      'filter': user_filter,
                      'alimento_1': alimento_1,
                      'alimento_2': alimento_2,
                      'alimento_3': alimento_3,
                      'alimento_4': alimento_4,
                      'alimento_6': alimento_6,
                      'alimento_7': alimento_7,
                      'alimento_8': alimento_8,
                      'alimento_9': alimento_9,
                      'alimento_10': alimento_10,
                      'alimento_11': alimento_11,
                      'alimento_12': alimento_12,
                      'alimento_13': alimento_13,
                      'alimento_14': alimento_14,
                      'alimento_15': alimento_15,
                      'nbar': "buscar_av"

                  })


def set_need_appearances_writer(writer):
    """
    Helper para escribir el pdf
    :param writer: el pdf
    :return:
    """
    # See 12.7.2 and 7.7.2 for more information:
    # http://www.adobe.com/content/dam/acom/en/devnet/acrobat/pdfs/PDF32000_2008.pdf
    try:
        catalog = writer._root_object
        # get the AcroForm tree and add "/NeedAppearances attribute
        if "/AcroForm" not in catalog:
            writer._root_object.update({
                NameObject("/AcroForm"): IndirectObject(len(writer._objects), 0, writer)})

        need_appearances = NameObject("/NeedAppearances")
        writer._root_object["/AcroForm"][need_appearances] = BooleanObject(
            True)
        return writer

    except Exception as e:
        print('set_need_appearances_writer() catch : ', repr(e))
        return writer


def generar_hoja_entrega(request, pk):
    """
    Generador de hoja de entrega para cada beneficiario en parte
    :param request:
    :param pk: id persona
    :return: pdf generado
    """

    dels = DelegacionData.objects.all().first()
    infile = os.path.join(settings.PROJECT_ROOT, 'entrega2020.pdf')
    inputStream = open(infile, "rb")
    pdf_reader = PdfFileReader(inputStream, strict=False)
    if "/AcroForm" in pdf_reader.trailer["/Root"]:
        pdf_reader.trailer["/Root"]["/AcroForm"].update(
            {NameObject("/NeedAppearances"): BooleanObject(True)})

    pdf_writer = PdfFileWriter()
    set_need_appearances_writer(pdf_writer)
    if "/AcroForm" in pdf_writer._root_object:
        pdf_writer._root_object["/AcroForm"].update(
            {NameObject("/NeedAppearances"): BooleanObject(True)})

    persona = Persona.objects.get(id=pk)
    familiares = persona.hijo.all()
    familiares_gr = persona.hijo.filter(edad__gt=3)

    mayores = 0
    menores = 0
    for f in familiares:
        if calculate_age(f.fecha_nacimiento) > 3:
            mayores += 1
        else:
            menores += 1

    field_dictionary = {
        "NombreOAR": f"{request.tenant}",
        "DireccioOAR": f"{dels.direccion}",
        "Nombre y apellidos del representante de la unidad familiar": f"{persona.nombre_apellido}",
        "DNINIEPasaporte 1": f"{persona.dni}",
        "Teléfono": f"{persona.telefono}",
        "Domicilio": f"{persona.domicilio}",
        "Localidad": f"{persona.ciudad}",
        "CP": "28850",
        "TOTAL MIEMBROS UNIDAD FAMILIAR": f"{mayores + menores + 1}",
        "Niños 02 ambos inclusive": f"{menores}",
        "numarAdra": f"{persona.numero_adra}"
    }

    pdf_writer.addPage(pdf_reader.getPage(0))
    pdf_writer.updatePageFormFieldValues(
        pdf_writer.getPage(0), field_dictionary)

    # outputStream = open(outfile, "wb")
    # pdf_writer.write(outputStream)

    # outputStream.close()

    # extractedPage = open(pdf_file_path, 'rb')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment;filename="{persona.numero_adra}.pdf"'
    pdf_writer.write(response)
    inputStream.close()
    return response


def generar_hoja_valoracion_social(request, pk):
    persona = Persona.objects.get(pk=pk, active=True)
    infile = os.path.join(settings.PROJECT_ROOT, 'vl.docx')
    template = infile
    document = MailMerge(template)

    hijos = []

    for h in persona.hijo.all():
        print(h.otros_documentos)
        hijo_dict = {}
        hijo_dict['parentesco'] = f'{h.parentesco}'
        hijo_dict['nombre_apellido_hijo'] = f'{h.nombre_apellido}'
        hijo_dict['dni_hijo'] = f'{h.dni}'
        hijo_dict['pas_hijo'] = f'{h.otros_documentos}'
        hijo_dict['fecha_nacimiento_hijo'] = f"{'{:%d-%m-%Y}'.format(h.fecha_nacimiento)}"
        hijos.append(hijo_dict)
    document.merge(
        numar_adra=f'{persona.numero_adra}',
        nombre_apellido=f'{persona.nombre_apellido}',
        dni=f'{persona.dni if persona.dni else persona.otros_documentos}',
        fecha_nacimiento=f"{'{:%d-%m-%Y}'.format(persona.fecha_nacimiento)}",
        nacionalidad=f'{persona.nacionalidad}',
        domicilio=f'{persona.domicilio}',
        ciudad=f'{persona.ciudad}',
        numar_telefon=f'{persona.telefono}',
        # fecha_hoy=f"{'{:%d-%m-%Y}'.format(persona.created_at) }",
        fecha_hoy="",

    )
    if persona.empadronamiento:
        document.merge(a="x")
    if persona.libro_familia:
        document.merge(b="x")
    if persona.fotocopia_dni:
        document.merge(c="x")
    if persona.prestaciones:
        document.merge(d="x")
    if persona.nomnia:
        document.merge(e="x")
    if persona.cert_negativo:
        document.merge(f="x")
    if persona.aquiler_hipoteca:
        document.merge(g="x")
    if persona.recibos:
        document.merge(h="x")
    document.merge_rows('parentesco', hijos)

    # document.write(f'./valoracion/{p.numero_adra}.docx')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename={persona.numero_adra}.docx'
    document.write(response)

    return response


# API ADRA
# @api_view(['GET', 'POST'])
# def persona_list(request,format=None):
#     """
#     Listado todas las personas
#     """
#     if request.method == 'GET':
#         personas = Persona.objects.all()
#         serializer = PersonaSerializer(personas, many=True)
#         return Response(serializer.data)
#
#     elif request.method == 'POST':
#         serializer = PersonaSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#
#
# @api_view(['GET', 'PUT', 'DELETE'])
# def persona_detail(request, pk,format=None):
#     """
#     Crud persona.
#     """
#     try:
#         persona = Persona.objects.get(pk=pk)
#     except Persona.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)
#
#     if request.method == 'GET':
#         serializer = PersonaSerializer(persona)
#         return Response(serializer.data)
#
#     elif request.method == 'PUT':
#         serializer = PersonaSerializer(persona, data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#
#     elif request.method == 'DELETE':
#         persona.delete()
#         return Response(status=status.HTTP_204_NO_CONTENT)


class PersonaViewSet(viewsets.ModelViewSet):
    """
        This viewset automatically provides `list`, `create`, `retrieve`,
        `update` and `destroy` actions.
        """
    queryset = Persona.objects.all().order_by("-numero_adra")
    serializer_class = PersonaSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'telefono'

    def perform_create(self, serializer):
        serializer.save(modificado_por=self.request.user)


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    """
    This viewset automatically provides `list` and `detail` actions.
    """

    queryset = get_user_model().objects.all()
    serializer_class = UserSerializer


def get_data(request):
    """" manda datos a los graficos de echats """
    list_ano = []
    registro_2018 = Persona.objects.filter(created_at__year=2018)
    list_ano.append(registro_2018.count())
    registro_2019 = Persona.objects.filter(created_at__year=2019)
    list_ano.append(registro_2019.count())
    registro_2020 = Persona.objects.filter(created_at__year=2020)
    list_ano.append(registro_2020.count())
    registro_2021 = Persona.objects.filter(created_at__year=2021)
    list_ano.append(registro_2021.count())

    return JsonResponse({"reg": list_ano})


class CustomAllauthAdapter(DefaultAccountAdapter):

    def send_mail(self, template_prefix, email, context):

        dels = DelegacionData.objects.all().first()
        hostname_without_port = remove_www(self.request.get_host().split(':')[0])

        if context.get('activate_url'):
            account_confirm_email = 'accounts/confirm-email/'
            context['activate_url'] = (
                        "http://" + str(hostname_without_port) + "/" + account_confirm_email + context['key'])
            user = context.get('user')

            sg = SendGridAPIClient(settings.SENDGRID_API_KEY)
            message = Mail(
                from_email=f"admin@adra.es",
                to_emails=f"{user.email}",
            )
            message.dynamic_template_data = {
                "activate_url": f"{context.get('activate_url')}",
                "user": f"{context.get('user')}",
                "Sender_Name": f"{dels.nombre}",
                "Sender_Address": f"{dels.direccion}",
                "Sender_City": f"{dels.ciudad}",
                "Sender_State": f"{dels.provincia}",
                "Sender_Zip": f"{dels.codigo_postal}"
            }
            message.template_id = 'd-8dddee085b5e4479a28b7dace0adf686'
            response = sg.send(message)

        elif context.get('password_reset_url'):

            user = context.get('user')
            sg = SendGridAPIClient(settings.SENDGRID_API_KEY)
            message = Mail(
                from_email=f"admin@adra.es",
                to_emails=f"{user.email}",
            )
            message.dynamic_template_data = {
                "url_cambiar": f"{context.get('password_reset_url')}",
                "user": f"{user}",
                "Sender_Name": f"{dels.nombre}",
                "Sender_Address": f"{dels.direccion}",
                "Sender_City": f"{dels.ciudad}",
                "Sender_State": f"{dels.provincia}",
                "Sender_Zip": f"{dels.codigo_postal}"
            }

            message.template_id = 'd-ab0adafe4dd14cb4b9aba688b7200830'
            response = sg.send(message)
